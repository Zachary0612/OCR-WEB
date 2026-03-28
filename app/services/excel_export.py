"""
归档文件目录 Excel 导出服务
从 OCR 结果中提取关键字段，写入/追加到 Excel
"""
import re
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

HEADERS = ['档号', '文号', '责任者', '题名', '日期', '页数', '密级', '备注']

DEFAULT_EXCEL_NAME = '归档文件目录.xlsx'


def extract_fields(filename: str, full_text: str, result_json, page_count: int) -> dict:
    """从 OCR 结果中提取关键字段"""
    fields = {h: "" for h in HEADERS}
    fields["页数"] = str(page_count) if page_count else ""

    if not full_text:
        full_text = ""

    text_clean = re.sub(r'\s+', ' ', full_text).strip()
    lines = [l.strip() for l in full_text.split('\n') if l.strip()]

    # --- 档号：从文件名提取 ---
    fname_stem = Path(filename).stem
    dh_match = re.match(r'(WS[·.]?\d{4}[·.]?[A-Z]\d+[-]\d+)', fname_stem)
    if dh_match:
        fields["档号"] = dh_match.group(1)
    elif re.match(r'(KJ[-].*)', fname_stem):
        parts = fname_stem.split('-')
        if len(parts) >= 5:
            fields["档号"] = '-'.join(parts[:5])

    # --- 文号 ---
    wh_patterns = [
        r'[\u4e00-\u9fa5]+[\[〔\(（]?\d{4}[\]〕\)）]?\s*(?:第\s*)?\d+\s*号',
        r'[\u4e00-\u9fa5]{2,10}发[\[〔\(（]\d{4}[\]〕\)）]\d+号',
        r'[\u4e00-\u9fa5]{2,10}[\[〔\(（]\d{4}[\]〕\)）]\s*\d+\s*号',
    ]
    for pat in wh_patterns:
        m = re.search(pat, text_clean)
        if m:
            fields["文号"] = m.group(0).strip()
            break

    # --- 题名：优先从 regions 标题类型 ---
    if result_json:
        pages = result_json if isinstance(result_json, list) else [result_json]
        for page in pages:
            if not isinstance(page, dict):
                continue
            for region in page.get("regions", []):
                rtype = region.get("type", "")
                content = region.get("content", "")
                if rtype in ("doc_title", "title", "paragraph_title", "content_title") and content:
                    if len(content) > len(fields["题名"]):
                        fields["题名"] = content.strip()

    if not fields["题名"]:
        for line in lines[:10]:
            if len(line) >= 6 and not re.match(r'^第?\d+页', line):
                if re.search(r'(关于|通知|决定|意见|办法|规则|方法|规范|条例|规定)', line):
                    fields["题名"] = line
                    break

    if not fields["题名"] and lines:
        candidates = sorted(lines[:8], key=len, reverse=True)
        if candidates:
            fields["题名"] = candidates[0][:100]

    # --- 责任者 ---
    resp_patterns = [
        r'([\u4e00-\u9fa5]{2,20}(?:局|部|委员会|委|办|厅|院|会|中心|处|科|室))\s*(?:关于|发布|印发)',
        r'([\u4e00-\u9fa5]{4,20}(?:人民政府|人力资源|档案馆|档案局))',
    ]
    for pat in resp_patterns:
        m = re.search(pat, text_clean)
        if m:
            fields["责任者"] = m.group(1).strip()
            break
    if not fields["责任者"] and fields["文号"]:
        m = re.match(r'([\u4e00-\u9fa5]+)', fields["文号"])
        if m:
            fields["责任者"] = m.group(1)

    # --- 日期 ---
    date_patterns = [
        r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日',
        r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})',
    ]
    for pat in date_patterns:
        m = re.search(pat, text_clean)
        if m:
            y, mo, d = m.group(1), m.group(2), m.group(3)
            fields["日期"] = f"{y}-{int(mo):02d}-{int(d):02d}"
            break

    # --- 密级 ---
    mj_match = re.search(r'(绝密|机密|秘密|内部|公开)', text_clean[:200])
    if mj_match:
        fields["密级"] = mj_match.group(1)

    return fields


def resolve_excel_output_path(output_path: str) -> str:
    raw = (output_path or '').strip()
    p = Path(raw)
    if raw.endswith(('\\', '/')) or (p.exists() and p.is_dir()) or not p.suffix:
        p = p / DEFAULT_EXCEL_NAME
    elif p.suffix.lower() == '.xls':
        p = p.with_suffix('.xlsx')
    p.parent.mkdir(parents=True, exist_ok=True)
    return str(p)


def init_excel(output_path: str) -> str:
    """创建 Excel 文件，写入表头，返回路径"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "归档文件目录"

    ws.merge_cells('A1:H1')
    ws['A1'] = '归档文件目录'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    col_widths = {'A': 20, 'B': 25, 'C': 15, 'D': 50, 'E': 12, 'F': 6, 'G': 8, 'H': 15}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    logger.info("创建归档目录 Excel: %s", output_path)
    return output_path


def clear_excel_data(output_path: str):
    """清空 Excel 数据行（保留标题行和表头行），用于每次批量写入前重置"""
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    # 删除第3行及以后所有数据行
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    wb.save(output_path)
    logger.info("已清空归档目录数据行: %s", output_path)


def append_to_excel(output_path: str, fields: dict):
    """向 Excel 追加一行数据"""
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    next_row = ws.max_row + 1
    for col_idx, header in enumerate(HEADERS, 1):
        ws.cell(row=next_row, column=col_idx, value=fields.get(header, ""))
    wb.save(output_path)
