"""
归档记录服务：保存 / 查询 / 导出 / 导入 archive_records 表
"""
import logging
import tempfile
import os
from pathlib import Path

import openpyxl
from sqlalchemy import select, func, delete as sa_delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import ArchiveRecord
from app.services.excel_export import HEADERS, init_excel, append_to_excel

logger = logging.getLogger(__name__)


async def save_archive_record(
    db: AsyncSession,
    task_id: int | None,
    batch_id: str,
    batch_folder: str,
    fields: dict,
) -> ArchiveRecord:
    """保存归档记录（按 task_id UPSERT；task_id 为 None 时直接插入）"""
    record: ArchiveRecord | None = None
    if task_id is not None:
        res = await db.execute(
            select(ArchiveRecord).where(ArchiveRecord.task_id == task_id)
        )
        record = res.scalar_one_or_none()

    if record is None:
        record = ArchiveRecord(task_id=task_id)
        db.add(record)

    record.batch_id = batch_id
    record.batch_folder = batch_folder
    record.archive_no = fields.get("档号", "") or ""
    record.doc_no = fields.get("文号", "") or ""
    record.responsible = fields.get("责任者", "") or ""
    record.title = fields.get("题名", "") or ""
    record.date = fields.get("日期", "") or ""
    record.pages = fields.get("页数", "") or ""
    record.classification = fields.get("密级", "") or ""
    record.remarks = fields.get("备注", "") or ""

    await db.commit()
    await db.refresh(record)
    return record


async def get_archive_records(
    db: AsyncSession,
    folder: str = "",
    batch_id: str = "",
    page: int = 1,
    page_size: int = 200,
):
    """查询归档记录列表，支持按 folder / batch_id 过滤"""
    q = select(ArchiveRecord)
    cnt_q = select(func.count()).select_from(ArchiveRecord)

    if folder:
        q = q.where(ArchiveRecord.batch_folder == folder)
        cnt_q = cnt_q.where(ArchiveRecord.batch_folder == folder)
    if batch_id:
        q = q.where(ArchiveRecord.batch_id == batch_id)
        cnt_q = cnt_q.where(ArchiveRecord.batch_id == batch_id)

    total = (await db.execute(cnt_q)).scalar() or 0
    q = q.order_by(ArchiveRecord.created_at.asc()).offset((page - 1) * page_size).limit(page_size)
    records = (await db.execute(q)).scalars().all()
    return records, total


def records_to_excel(records: list, output_path: str) -> str:
    """将归档记录列表写入 Excel（先 init 再 append）"""
    init_excel(output_path)
    for rec in records:
        fields = {
            "档号": rec.archive_no or "",
            "文号": rec.doc_no or "",
            "责任者": rec.responsible or "",
            "题名": rec.title or "",
            "日期": rec.date or "",
            "页数": rec.pages or "",
            "密级": rec.classification or "",
            "备注": rec.remarks or "",
        }
        append_to_excel(output_path, fields)
    logger.info("导出归档记录 %d 条 -> %s", len(records), output_path)
    return output_path


async def import_from_excel(db: AsyncSession, file_path: str, batch_id: str = "") -> int:
    """从 .xlsx / .xls 文件批量导入归档记录到数据库，返回导入行数"""
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ext = p.suffix.lower()
    rows: list[dict] = []

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        header_row_idx = None
        headers: list[str] = []
        for r in range(1, min(6, ws.max_row + 1)):
            vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 10)]
            if "档号" in vals or "文号" in vals:
                header_row_idx = r
                headers = vals
                break
        if header_row_idx is None:
            raise ValueError("未找到表头行（含'档号'或'文号'列）")
        for r in range(header_row_idx + 1, ws.max_row + 1):
            row = {
                headers[c - 1]: str(ws.cell(r, c).value or "").strip()
                for c in range(1, min(len(headers) + 1, 10))
            }
            if any(v for v in row.values()):
                rows.append(row)

    elif ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ImportError("读取 .xls 文件需要 xlrd，请运行: pip install xlrd==1.2.0")
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        header_row_idx = None
        headers = []
        for r in range(min(6, ws.nrows)):
            vals = [str(ws.cell_value(r, c)).strip() for c in range(min(9, ws.ncols))]
            if "档号" in vals or "文号" in vals:
                header_row_idx = r
                headers = vals
                break
        if header_row_idx is None:
            raise ValueError("未找到表头行（含'档号'或'文号'列）")
        for r in range(header_row_idx + 1, ws.nrows):
            row = {
                headers[c]: str(ws.cell_value(r, c)).strip()
                for c in range(min(len(headers), ws.ncols))
            }
            if any(v for v in row.values()):
                rows.append(row)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx 或 .xls")

    count = 0
    folder = str(p.parent)
    for row in rows:
        record = ArchiveRecord(
            task_id=None,
            batch_id=batch_id or f"import_{p.stem}",
            batch_folder=folder,
            archive_no=row.get("档号", ""),
            doc_no=row.get("文号", ""),
            responsible=row.get("责任者", ""),
            title=row.get("题名", ""),
            date=row.get("日期", ""),
            pages=row.get("页数", ""),
            classification=row.get("密级", ""),
            remarks=row.get("备注", ""),
        )
        db.add(record)
        count += 1

    await db.commit()
    logger.info("从 %s 导入归档记录 %d 条", file_path, count)
    return count
